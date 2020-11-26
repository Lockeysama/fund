import os
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


def _make_headers(sheet, headers):
    cell_range = headers.get('range')
    sheet.merge_cells(cell_range)
    coordinates = cell_range.split(':')[0]
    c, r = re.sub(r'\d', '', coordinates), re.sub(r'\D', '', coordinates)
    c, r = ord(c) + 1 - 65, int(r)
    cell = sheet.cell(r, c, headers.get('title'))
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(bold=True)
    for index, sub in enumerate(headers.get('subs', [])):
        sub_c, sub_r = c + index, r + 1
        sub_cell = sheet.cell(sub_r, sub_c, sub.get('title'))
        sub_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sub_cell.font = Font(bold=True)


def _get_index(sheet, headers):
    sub_num = len(headers.get('subs'))
    start = 1
    while start < 3000:
        base = chr(64 + (start // 26)) if start // 26 else ''
        x1 = chr(65 + start % 26)
        x2 = chr(65 + start % 26 + sub_num - 1)
        title_range = '{}{}{}:{}{}{}'.format(
            base, x1, len(headers.get('items')) + 1,
            base, x2, len(headers.get('items')) + 1
        )
        title = sheet[title_range][0][0].value
        if title == headers.get('title'):
            return False, start, '{}{}'.format(base, x1)
        if not title:
            headers['range'] = title_range
            _make_headers(sheet, headers)
            return True, start, '{}{}'.format(base, x1)
        start += sub_num
    return False, -1,


def _insert_data(sheet, headers, column, all_data):
    header_data, data = all_data
    row_date = datetime.strptime(data[0][0], '%Y-%m-%d')
    first_row = sheet['A{}'.format(len(headers.get('items')) + 3)].value
    base = chr(64 + (column // 26)) if column // 26 else ''
    x1 = chr(65 + column % 26)
    x2 = chr(65 + column % 26 + len(headers.get('subs')) - 1)
    for index, d in enumerate(header_data):
        column_range = '{}{}{}:{}{}{}'.format(
            base, x1, index + 1,
            base, x2, index + 1
        )
        sheet.merge_cells(column_range)
        sheet.cell(index + 1, column + 1, d)
    if first_row:
        first_row_date = datetime.strptime(first_row, '%Y-%m-%d')
        nr = len(headers.get('items')) + 3
        for index, day in enumerate(range(0, (row_date - first_row_date).days, 7)):
            new_row_date = datetime.fromtimestamp(
                first_row_date.timestamp() + 86400 * (day + 7)
            ).strftime('%Y-%m-%d')
            sheet.insert_rows(nr)
            sheet['A{}'.format(nr)] = new_row_date
    else:
        for index in range(100):
            new_row_date = datetime.fromtimestamp(
                row_date.timestamp() - (86400 * (index * 7))
            ).strftime('%Y-%m-%d')
            sheet['A{}'.format(index + len(headers.get('items')) + 3)] = new_row_date
    for item in reversed(data):
        for index in range(len(headers.get('items')) + 3, 1000):
            table_date = sheet['A{}'.format(index)].value
            try:
                table_date = datetime.strptime(table_date, '%Y-%m-%d')
            except Exception as e:
                print(e)
            row_date = datetime.strptime(item[0], '%Y-%m-%d')
            if row_date == table_date:
                for i, v in enumerate(item[1:]):
                    try:
                        sheet.cell(index, column + i + 1, v)
                    except Exception as e:
                        print(e)
                break


def make(tactics, headers, data):
    file_name = './fond.xlsx'
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    try:
        sheet = wb.get_sheet_by_name(tactics)
    except KeyError as e:
        sheet = wb.create_sheet(tactics)
        for index, item in enumerate(headers.get('items')):
            cell = sheet.cell(index + 1, 1, item)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
        cell_range = 'A{}:A{}'.format(len(headers.get('items')) + 1, len(headers.get('items')) + 2)
        sheet.merge_cells(cell_range)
        sheet.cell(len(headers.get('items')) + 1, 1, '名称')
    new_column, column, column_range = _get_index(sheet, headers)
    _insert_data(sheet, headers, column, data)
    wb.save(file_name)


if __name__ == '__main__':
    a = [[], [['2020-11-20', '9.0420', '8.9320'], ['2020-11-13', '8.0420', '8.9320'], ['2020-11-06', '7.9150', '8.8050'], ['2020-10-30', '7.7691', '8.6591'], ['2020-10-23', '7.5531', '8.4431']]]
    b = [[], [['2020-12-04', '7.6666', '5.8888'], ['2020-11-27', '7.6666', '5.8888'], ['2020-11-20', '3.6666', '5.8888'], ['2020-11-13', '3.6620', '5.3590'], ['2020-11-06', '3.6570', '5.3540'], ['2020-10-30', '3.5060', '5.2030'], ['2020-10-23', '3.4860', '5.1830']]]
    c = [[], [['2020-11-20', '22.6666', '22.8888'], ['2020-11-13', '3.6620', '5.3590'], ['2020-11-06', '3.6570', '5.3540'],
         ['2020-10-30', '3.5060', '5.2030'], ['2020-10-23', '3.4860', '5.1830']]]
    d = [[], [['2020-11-06', '7.9150', '8.8050'], ['2020-10-30', '7.7691', '8.6591'],
         ['2020-10-23', '7.5531', '8.4431'], ['2020-10-16', '66.5531', '8.4431'], ['2020-09-25', '66.5531', '8.4431']]]
    e = [[], [['2020-11-06', '7.9150', '8.8050'], ['2020-10-30', '7.7691', '8.6591'],
         ['2020-10-23', '7.5531', '8.4431'], ['2020-10-02', '66.5531', '8.4431'], ['2020-09-25', '66.5531', '8.4431']]]
    data = [
        [110011, a],
        [519069, b],
        [519068, c],
        [519099, a],
        [519078, d],
        [522099, e],
    ]
    increase = ['阶段涨幅', '同类平均', '沪深 300', '同类排名']
    cycle = ['近 1 周', '近 1 月', '近 3 月', '近 6 月', '近 1 年']
    increase_cycle = ['{}（{}）'.format(i, c) for i in increase for c in cycle]
    for fid, d in data:
        headers = {
            'items': increase_cycle,
            'title': '易方达（{}）'.format(fid),
            'subs': [
                {'title': '单位净值'},
                {'title': '累计净值'},
            ]
        }
        make('公募', headers, d)
